"""Batch processing input template generator.

Generates downloadable Excel or CSV templates that users fill in
with site data for batch analysis.
"""

from __future__ import annotations

import csv
import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

TEMPLATE_COLUMNS: list[str] = [
    "name",
    "latitude",
    "longitude",
    "analysis_type",
    "racking",
    "gcr",
    "dc_capacity_mw",
    "ac_capacity_mw",
    "module",
    "inverter",
    "tilt",
    "azimuth",
    "bifacial",
    "buildable_acres",
    "utilization_factor",
    "gcr_min",
    "gcr_max",
    "dcac_min",
    "dcac_max",
    "solar_cost_per_kw_dc",
    "solar_opex_per_kw_dc_year",
    "discount_rate_pct",
    "project_lifetime_years",
    "degradation_pct",
    "itc_pct",
    "energy_price_per_kwh",
    "energy_cost_escalator_pct",
    "dispatch_mode",
    "charging_mode",
    "bess_cost_per_kwh",
    "bess_opex_per_kw_year",
    "shading_pct",
    "dc_wiring_pct",
    "ac_wiring_pct",
    "transformer_pct",
    "availability_pct",
    "mismatch_pct",
    "lid_pct",
    "soiling_pct",
]

COLUMN_DESCRIPTIONS: dict[str, str] = {
    "name": "* Site name (unique identifier)",
    "latitude": "* Latitude (-90 to 90)",
    "longitude": "* Longitude (-180 to 180)",
    "analysis_type": "* buildability | production | optimization | optimization_bess",
    "racking": "fixed | tracker (default: tracker)",
    "gcr": "Ground coverage ratio 0-1 (production only)",
    "dc_capacity_mw": "DC nameplate capacity in MW (production only)",
    "ac_capacity_mw": "AC inverter capacity in MW (production only)",
    "module": "Exact CEC module name (default: LONGi LR5-72HBD-550M)",
    "inverter": "Exact CEC inverter name (default: Sungrow SG250HX-US [800V])",
    "tilt": "Tilt angle 0-90 (default: 60 tracker, 25 fixed)",
    "azimuth": "Azimuth 0-360 (default: 180 = south)",
    "bifacial": "TRUE or FALSE (default: FALSE)",
    "buildable_acres": "Buildable land area in acres (optimization only)",
    "utilization_factor": "Land utilization 0-1 (default: 0.85)",
    "gcr_min": "GCR sweep minimum (optimization only)",
    "gcr_max": "GCR sweep maximum (optimization only)",
    "dcac_min": "DC:AC ratio sweep minimum (default: 1.15)",
    "dcac_max": "DC:AC ratio sweep maximum (default: 1.60)",
    "solar_cost_per_kw_dc": "Solar installed cost $/kW-DC",
    "solar_opex_per_kw_dc_year": "Solar O&M $/kW-DC/year",
    "discount_rate_pct": "Discount rate %",
    "project_lifetime_years": "Project lifetime in years",
    "degradation_pct": "Annual degradation % (default: 0.5)",
    "itc_pct": "Investment tax credit %",
    "energy_price_per_kwh": "Energy price $/kWh (for NPV mode)",
    "energy_cost_escalator_pct": "Annual energy price escalation %",
    "dispatch_mode": "ftm (front-of-meter only in batch v1)",
    "charging_mode": "solar_and_grid | solar_only (default: solar_and_grid)",
    "bess_cost_per_kwh": "BESS installed cost $/kWh",
    "bess_opex_per_kw_year": "BESS O&M $/kW/year",
    "shading_pct": "Shading loss % (default: 0.0)",
    "dc_wiring_pct": "DC wiring loss % (default: 2.0)",
    "ac_wiring_pct": "AC wiring loss % (default: 0.5)",
    "transformer_pct": "Transformer loss % (default: 1.0)",
    "availability_pct": "Availability loss % (default: 2.5)",
    "mismatch_pct": "Module mismatch % (default: 2.0)",
    "lid_pct": "Light-induced degradation % (default: 1.5)",
    "soiling_pct": "Soiling loss % (default: 2.0)",
}

EXAMPLE_ROWS: list[dict[str, object]] = [
    {
        "name": "Phoenix South",
        "latitude": 33.45,
        "longitude": -112.07,
        "analysis_type": "buildability",
    },
    {
        "name": "Denver East",
        "latitude": 39.74,
        "longitude": -104.99,
        "analysis_type": "production",
        "racking": "tracker",
        "gcr": 0.40,
        "dc_capacity_mw": 5.0,
        "ac_capacity_mw": 4.0,
    },
    {
        "name": "Austin West",
        "latitude": 30.27,
        "longitude": -97.74,
        "analysis_type": "optimization",
        "buildable_acres": 50.0,
    },
    {
        "name": "PJM Site Alpha",
        "latitude": 39.95,
        "longitude": -75.16,
        "analysis_type": "optimization_bess",
        "buildable_acres": 80.0,
        "dispatch_mode": "ftm",
    },
]

INSTRUCTIONS_TEXT: str = """BATCH TEMPLATE INSTRUCTIONS
============================

OVERVIEW
--------
Fill in the "Template" sheet with up to 25 sites. Each row is one analysis.
Upload the completed file (CSV or Excel) to run all analyses in a single batch.

ANALYSIS TYPES
--------------
Each row must have an analysis_type. Required fields vary by type:

  buildability
    Required: name, latitude, longitude
    All other fields are ignored.

  production
    Required: name, latitude, longitude, racking, gcr, dc_capacity_mw, ac_capacity_mw
    Optional: module, inverter, tilt, azimuth, bifacial, loss parameters

  optimization
    Required: name, latitude, longitude, buildable_acres
    Optional: racking, module, inverter, gcr_min, gcr_max, dcac_min, dcac_max,
              utilization_factor, economic parameters (for LCOE/NPV mode)

  optimization_bess
    Required: name, latitude, longitude, buildable_acres, dispatch_mode
    Optional: Same as optimization, plus charging_mode, bess_cost_per_kwh,
              bess_opex_per_kw_year

EQUIPMENT DEFAULTS
------------------
If module or inverter is left blank, the system uses:
  Module:   LONGi Green Energy Technology Co. Ltd. LR5-72HBD-550M
  Inverter: Sungrow Power Supply Co - Ltd : SG250HX-US [800V]

Use the equipment search endpoint or chat to find exact CEC names.

LOSS PARAMETER DEFAULTS
------------------------
If any loss field is left blank, these defaults apply:
  shading_pct:      0.0
  dc_wiring_pct:    2.0
  ac_wiring_pct:    0.5
  transformer_pct:  1.0
  degradation_pct:  0.5
  availability_pct: 2.5
  mismatch_pct:     2.0
  lid_pct:          1.5
  soiling_pct:      2.0

LIMITS
------
- Maximum 25 rows per batch.
- Site names must be unique within the batch.

BESS NOTE
---------
Batch v1 supports front-of-meter (FTM) BESS dispatch only.
Behind-the-meter (BTM) is not supported in batch mode because it requires
per-site rate schedules and load profiles that cannot be specified in a
flat table.

FILE FORMAT
-----------
Both CSV (.csv) and Excel (.xlsx) uploads are accepted.
CSV files should use the same column headers as the Template sheet.
"""


def generate_batch_template(output_path: Path, format: str = "xlsx") -> Path:
    """Generate a batch input template file.

    Args:
        output_path: Directory or file path for the output. If a directory,
            the filename is auto-generated as ``batch_template.{format}``.
        format: Output format — ``"xlsx"`` or ``"csv"``.

    Returns:
        Path to the generated template file.

    Raises:
        ValueError: If format is not ``"xlsx"`` or ``"csv"``.
    """
    if format not in ("xlsx", "csv"):
        raise ValueError(f"Unsupported format: {format!r}. Use 'xlsx' or 'csv'.")

    output_path = Path(output_path)
    if output_path.is_dir():
        output_path = output_path / f"batch_template.{format}"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    if format == "xlsx":
        _generate_xlsx(output_path)
    else:
        _generate_csv(output_path)

    logger.info("Generated batch template: %s", output_path)
    return output_path


def _generate_xlsx(output_path: Path) -> None:
    """Generate an Excel batch template with Template and Instructions sheets."""
    wb = Workbook()

    # ── Sheet 1: Template ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Template"

    header_font = Font(bold=True)
    desc_font = Font(italic=True, color="666666")
    desc_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Row 1: Headers
    for col_idx, header in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font

    # Row 2: Descriptions
    for col_idx, header in enumerate(TEMPLATE_COLUMNS, start=1):
        desc = COLUMN_DESCRIPTIONS.get(header, "")
        cell = ws.cell(row=2, column=col_idx, value=desc)
        cell.font = desc_font
        cell.fill = desc_fill
        cell.alignment = Alignment(wrap_text=True)

    # Rows 3-6: Example data
    for row_offset, example in enumerate(EXAMPLE_ROWS):
        row_num = 3 + row_offset
        for col_idx, header in enumerate(TEMPLATE_COLUMNS, start=1):
            value = example.get(header)
            if value is not None:
                ws.cell(row=row_num, column=col_idx, value=value)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-size columns to fit headers and descriptions
    for col_idx, header in enumerate(TEMPLATE_COLUMNS, start=1):
        desc = COLUMN_DESCRIPTIONS.get(header, "")
        max_len = max(len(header), len(desc))
        # Cap at 40 to avoid absurdly wide columns
        width = min(max_len + 2, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Sheet 2: Instructions ──────────────────────────────────────────
    ws_instr = wb.create_sheet(title="Instructions")
    ws_instr.column_dimensions["A"].width = 90
    for line_idx, line in enumerate(INSTRUCTIONS_TEXT.strip().split("\n"), start=1):
        cell = ws_instr.cell(row=line_idx, column=1, value=line)
        # Bold the section headers (lines that are all-caps or underlines)
        if line.strip() and line.strip() == line.strip().upper() and line.strip().isalpha():
            cell.font = Font(bold=True, size=12)
        elif line.startswith("===") or line.startswith("---"):
            continue  # Leave underline rows as-is (text separators)

    wb.save(output_path)


def _generate_csv(output_path: Path) -> None:
    """Generate a CSV batch template with header row only."""
    with open(output_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(TEMPLATE_COLUMNS)
