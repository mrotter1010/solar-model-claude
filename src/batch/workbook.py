"""Batch results Excel workbook generator.

Takes a BatchResult and the original input rows and produces a formatted
multi-tab Excel workbook with per-analysis-type result sheets, an input
echo, and an errors sheet.
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.batch.executor import BatchResult, RowResult

logger = logging.getLogger(__name__)

# ── Tab ordering and display names ─────────────────────────────────────

# analysis_type value → display name for the tab
_TAB_DISPLAY_NAMES: dict[str, str] = {
    "buildability": "Buildability",
    "production": "Production",
    "optimization": "Optimization",
    "optimization_bess": "Optimization BESS",
}

# Canonical order for analysis tabs
_TAB_ORDER: list[str] = [
    "buildability",
    "production",
    "optimization",
    "optimization_bess",
]

# ── Column definitions per analysis type ───────────��───────────────────

# Common columns prepended to every analysis tab
_COMMON_COLUMNS: list[tuple[str, str]] = [
    ("name", "Site Name"),
    ("latitude", "Latitude"),
    ("longitude", "Longitude"),
    ("status", "Status"),
    ("error", "Error"),
    ("runtime_seconds", "Runtime (s)"),
]

# Analysis-specific result columns: (result_dict_key, header_label)
_BUILDABILITY_COLUMNS: list[tuple[str, str]] = [
    ("buildable_acres", "Buildable Acres"),
    ("buildable_pct", "Buildable %"),
    ("total_acres", "Total Acres"),
    ("dominant_land_cover", "Dominant Land Cover"),
    ("mean_slope", "Mean Slope (\u00b0)"),
    ("tracker_suitable_pct", "Tracker Suitable %"),
]

_PRODUCTION_COLUMNS: list[tuple[str, str]] = [
    ("annual_energy_mwh", "Annual Energy (MWh)"),
    ("capacity_factor_ac_pct", "Capacity Factor AC"),
    ("specific_yield_kwh_per_kw", "Specific Yield (kWh/kW)"),
    ("performance_ratio_pct", "Performance Ratio"),
    ("dc_capacity_mw", "DC Capacity (MW)"),
    ("ac_capacity_mw", "AC Capacity (MW)"),
    ("gcr", "GCR"),
    ("dcac_ratio", "DC:AC Ratio"),
]

_OPTIMIZATION_COLUMNS: list[tuple[str, str]] = [
    ("winner_type", "Winner Type"),
    ("winner_gcr", "Winner GCR"),
    ("winner_dcac", "Winner DC:AC"),
    ("winner_mw_dc", "Winner MW DC"),
    ("winner_mw_ac", "Winner MW AC"),
    ("winner_annual_mwh", "Winner Annual MWh"),
    ("winner_cf_pct", "Winner CF"),
    ("winner_lcoe_per_mwh", "Winner LCOE ($/MWh)"),
    ("winner_npv", "Winner NPV ($)"),
    ("buildable_acres", "Buildable Acres"),
]

_OPTIMIZATION_BESS_COLUMNS: list[tuple[str, str]] = [
    *_OPTIMIZATION_COLUMNS,
    ("bess_adds_value", "BESS Adds Value"),
    ("bess_power_mw", "BESS Power (MW)"),
    ("bess_duration_hr", "BESS Duration (hr)"),
    ("bess_npv", "BESS NPV ($)"),
    ("combined_npv", "Combined NPV ($)"),
]

_COLUMNS_BY_TYPE: dict[str, list[tuple[str, str]]] = {
    "buildability": _BUILDABILITY_COLUMNS,
    "production": _PRODUCTION_COLUMNS,
    "optimization": _OPTIMIZATION_COLUMNS,
    "optimization_bess": _OPTIMIZATION_BESS_COLUMNS,
}

# ── Number format rules ──────────────────���─────────────────────────────

# result_dict_key → Excel number format string
_NUMBER_FORMATS: dict[str, str] = {
    # MWh — commas, no decimals
    "annual_energy_mwh": '#,##0',
    "winner_annual_mwh": '#,##0',
    # Percentages — 1 decimal + %
    "buildable_pct": '0.0"%"',
    "tracker_suitable_pct": '0.0"%"',
    "capacity_factor_ac_pct": '0.0"%"',
    "performance_ratio_pct": '0.0"%"',
    "winner_cf_pct": '0.0"%"',
    # Dollars — 2 decimals
    "winner_lcoe_per_mwh": '#,##0.00',
    "winner_npv": '#,##0.00',
    "bess_npv": '#,##0.00',
    "combined_npv": '#,##0.00',
    # GCR / DC:AC — 2 decimals
    "gcr": '0.00',
    "dcac_ratio": '0.00',
    "winner_gcr": '0.00',
    "winner_dcac": '0.00',
    # Acres �� 1 decimal
    "buildable_acres": '#,##0.0',
    "total_acres": '#,##0.0',
    # Slope — 1 decimal
    "mean_slope": '0.0',
    # MW — 2 decimals
    "dc_capacity_mw": '#,##0.00',
    "ac_capacity_mw": '#,##0.00',
    "winner_mw_dc": '#,##0.00',
    "winner_mw_ac": '#,##0.00',
    "bess_power_mw": '#,##0.00',
    # Hours — 1 decimal
    "bess_duration_hr": '0.0',
    # Specific yield — 0 decimals
    "specific_yield_kwh_per_kw": '#,##0',
    # Runtime
    "runtime_seconds": '0.00',
}

# ── Style constants ────────────────��───────────────────────────────────

_HEADER_FILL_BLUE = PatternFill(
    start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"
)
_HEADER_FILL_RED = PatternFill(
    start_color="FADBD8", end_color="FADBD8", fill_type="solid"
)
_HEADER_FONT = Font(name="Calibri", size=11, bold=True)
_DEFAULT_FONT = Font(name="Calibri", size=11)
_ALT_ROW_FILL = PatternFill(
    start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
)

_TAB_COLOR_BLUE = "3498DB"
_TAB_COLOR_GRAY = "95A5A6"
_TAB_COLOR_RED = "E74C3C"

_MIN_COL_WIDTH = 10
_MAX_COL_WIDTH = 40


# ── Public API ─────────���───────────────��───────────────────────────────


def generate_batch_workbook(
    batch_result: BatchResult,
    input_rows: list[dict],
    output_path: Path | None = None,
) -> Path:
    """Generate a formatted Excel workbook from batch results.

    Creates one tab per analysis type present in the results, an Input Echo
    tab with the original input data, and an Errors tab if any rows failed.

    Args:
        batch_result: Completed BatchResult from the executor.
        input_rows: Original validated input row dicts (for the Input Echo tab).
        output_path: Where to save the workbook. If None, saves to
            ``batch_result.output_dir / "batch_results.xlsx"``.

    Returns:
        Path to the saved Excel file.
    """
    if output_path is None:
        output_path = batch_result.output_dir / "batch_results.xlsx"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove the default sheet — we'll create our own
    wb.remove(wb.active)

    # Group row results by analysis type
    rows_by_type: dict[str, list[RowResult]] = {}
    for rr in batch_result.row_results:
        rows_by_type.setdefault(rr.analysis_type, []).append(rr)

    # Create analysis tabs in canonical order
    for analysis_type in _TAB_ORDER:
        if analysis_type not in rows_by_type:
            continue
        tab_name = _TAB_DISPLAY_NAMES.get(analysis_type, analysis_type)
        result_columns = _COLUMNS_BY_TYPE.get(analysis_type, [])
        _write_analysis_tab(
            wb, tab_name, rows_by_type[analysis_type], result_columns, input_rows
        )

    # Input Echo tab
    _write_input_echo_tab(wb, input_rows)

    # Errors tab (only if any failures)
    failed = [rr for rr in batch_result.row_results if rr.status == "failed"]
    if failed:
        _write_errors_tab(wb, failed)

    wb.save(output_path)
    logger.info("Batch workbook saved: %s", output_path)
    return output_path


# ── Tab builders ───────────────────────────────────���───────────────────


def _write_analysis_tab(
    wb: Workbook,
    tab_name: str,
    row_results: list[RowResult],
    result_columns: list[tuple[str, str]],
    input_rows: list[dict],
) -> None:
    """Write a single analysis-type tab."""
    ws: Worksheet = wb.create_sheet(title=tab_name)
    ws.sheet_properties.tabColor = _TAB_COLOR_BLUE

    all_columns = _COMMON_COLUMNS + result_columns
    headers = [label for _, label in all_columns]
    keys = [key for key, _ in all_columns]

    # Build an index from name → input row for lat/lon lookup
    input_lookup: dict[str, dict] = {}
    for row in input_rows:
        name = row.get("name")
        if name is not None:
            input_lookup[str(name)] = row

    # Header row
    _write_header_row(ws, headers, _HEADER_FILL_BLUE)

    # Data rows
    for row_idx, rr in enumerate(row_results, start=2):
        input_row = input_lookup.get(rr.name, {})
        for col_idx, key in enumerate(keys, start=1):
            value = _get_cell_value(rr, input_row, key)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _DEFAULT_FONT
            # Number format
            fmt = _NUMBER_FORMATS.get(key)
            if fmt and value is not None:
                cell.number_format = fmt
        # Alternating row shading
        if row_idx % 2 == 0:
            _apply_row_fill(ws, row_idx, len(keys), _ALT_ROW_FILL)

    _finalize_sheet(ws, headers)


def _write_input_echo_tab(wb: Workbook, input_rows: list[dict]) -> None:
    """Write the Input Echo tab with raw input data."""
    ws: Worksheet = wb.create_sheet(title="Input Echo")
    ws.sheet_properties.tabColor = _TAB_COLOR_GRAY

    if not input_rows:
        _write_header_row(ws, ["(no input data)"], _HEADER_FILL_BLUE)
        _finalize_sheet(ws, ["(no input data)"])
        return

    # Collect all unique columns across all rows, preserving first-seen order
    all_keys: list[str] = []
    seen: set[str] = set()
    for row in input_rows:
        for key in row:
            if key not in seen:
                seen.add(key)
                all_keys.append(key)

    # Header row
    _write_header_row(ws, all_keys, _HEADER_FILL_BLUE)

    # Data rows
    for row_idx, row in enumerate(input_rows, start=2):
        for col_idx, key in enumerate(all_keys, start=1):
            value = row.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _DEFAULT_FONT
        if row_idx % 2 == 0:
            _apply_row_fill(ws, row_idx, len(all_keys), _ALT_ROW_FILL)

    _finalize_sheet(ws, all_keys)


def _write_errors_tab(wb: Workbook, failed: list[RowResult]) -> None:
    """Write the Errors tab listing all failed rows."""
    ws: Worksheet = wb.create_sheet(title="Errors")
    ws.sheet_properties.tabColor = _TAB_COLOR_RED

    headers = ["row_number", "name", "analysis_type", "status", "error_message"]

    _write_header_row(ws, headers, _HEADER_FILL_RED)

    for row_idx, rr in enumerate(failed, start=2):
        values = [rr.row_index, rr.name, rr.analysis_type, rr.status, rr.error]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _DEFAULT_FONT
        if row_idx % 2 == 0:
            _apply_row_fill(ws, row_idx, len(headers), _ALT_ROW_FILL)

    _finalize_sheet(ws, headers)


# ── Helpers ────────────────��────────────────────────────────��──────────


def _get_cell_value(rr: RowResult, input_row: dict, key: str) -> object:
    """Resolve a cell value from a RowResult + its input row.

    Common columns (name, lat, lon, status, error, runtime) are pulled
    from the RowResult or input row. Analysis-specific columns come
    from rr.results dict.
    """
    if key == "name":
        return rr.name
    if key == "latitude":
        return input_row.get("latitude")
    if key == "longitude":
        return input_row.get("longitude")
    if key == "status":
        return rr.status
    if key == "error":
        return rr.error
    if key == "runtime_seconds":
        return rr.runtime_seconds

    # Analysis-specific columns
    if rr.results is not None:
        return rr.results.get(key)
    return None


def _write_header_row(
    ws: Worksheet, headers: list[str], fill: PatternFill
) -> None:
    """Write and style the header row (row 1)."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _apply_row_fill(
    ws: Worksheet, row_idx: int, num_cols: int, fill: PatternFill
) -> None:
    """Apply a fill to every cell in a row."""
    for col_idx in range(1, num_cols + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill


def _finalize_sheet(ws: Worksheet, headers: list[str]) -> None:
    """Apply freeze panes, auto-filter, and auto-size columns."""
    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    if headers:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}1"

    # Auto-size columns
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        # Sample up to 50 data rows for width estimation
        for row_idx in range(2, min(ws.max_row + 1, 52)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_len = max(max_len, len(str(cell_value)))
        width = min(max(max_len + 2, _MIN_COL_WIDTH), _MAX_COL_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
