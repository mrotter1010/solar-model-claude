"""Batch input file validation.

Parses uploaded CSV/Excel files and validates all rows before
any analysis runs. Collects all errors across all rows (not fail-fast).
"""

from __future__ import annotations

import logging
import math
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd

from src.batch.template import TEMPLATE_COLUMNS
from src.pysam_integration.cec_database import CECDatabase
from src.pysam_integration.exceptions import (
    InverterNotFoundError,
    ModuleNotFoundError,
)

logger = logging.getLogger(__name__)

MAX_BATCH_ROWS = 25

VALID_ANALYSIS_TYPES = frozenset({
    "buildability",
    "production",
    "optimization",
    "optimization_bess",
})

REQUIRED_FIELDS_ALL = ("name", "latitude", "longitude", "analysis_type")

REQUIRED_FIELDS_BY_TYPE: dict[str, tuple[str, ...]] = {
    "buildability": (),
    "production": ("racking", "gcr", "dc_capacity_mw", "ac_capacity_mw"),
    "optimization": ("buildable_acres",),
    "optimization_bess": ("buildable_acres", "dispatch_mode"),
}

VALID_RACKING = frozenset({"tracker", "fixed"})
VALID_DISPATCH_MODE = frozenset({"btm", "ftm"})
VALID_CHARGING_MODE = frozenset({"solar_only", "solar_and_grid"})

# Columns whose values are percentages (0-100)
PCT_COLUMNS = frozenset({
    "shading_pct",
    "dc_wiring_pct",
    "ac_wiring_pct",
    "transformer_pct",
    "availability_pct",
    "mismatch_pct",
    "lid_pct",
    "soiling_pct",
    "discount_rate_pct",
    "degradation_pct",
    "itc_pct",
    "energy_cost_escalator_pct",
})

# Columns whose values are costs/prices (>= 0)
COST_COLUMNS = frozenset({
    "solar_cost_per_kw_dc",
    "solar_opex_per_kw_dc_year",
    "energy_price_per_kwh",
    "bess_cost_per_kwh",
    "bess_opex_per_kw_year",
})


@dataclass
class ValidationError:
    """A single validation error tied to a specific row and column."""

    row: int
    column: str
    value: Any
    message: str


@dataclass
class ValidationResult:
    """Aggregated result of batch input validation."""

    valid_rows: list[dict] = field(default_factory=list)
    errors: list[ValidationError] = field(default_factory=list)
    warnings: list[ValidationError] = field(default_factory=list)
    row_count: int = 0
    is_valid: bool = False


def validate_batch_input(
    file_path: Path,
    cec_db: CECDatabase | None = None,
) -> ValidationResult:
    """Validate a batch input CSV or Excel file.

    Reads the file, checks all rows for required fields, valid types,
    valid ranges, and CEC equipment names. Collects all errors before
    returning.

    Args:
        file_path: Path to a ``.csv`` or ``.xlsx`` batch input file.
        cec_db: Optional pre-loaded CEC database. If ``None``, a new
            instance is created.

    Returns:
        ValidationResult with valid rows, errors, and warnings.

    Raises:
        ValueError: If the file extension is not ``.csv`` or ``.xlsx``.
    """
    file_path = Path(file_path)
    ext = file_path.suffix.lower()
    if ext not in (".csv", ".xlsx"):
        raise ValueError(
            f"Unsupported file extension: {ext!r}. Use '.csv' or '.xlsx'."
        )

    df = _read_file(file_path, ext)

    result = ValidationResult(row_count=len(df))

    if df.empty:
        result.errors.append(
            ValidationError(row=0, column="", value=None, message="File contains no data rows.")
        )
        return result

    if len(df) > MAX_BATCH_ROWS:
        result.errors.append(
            ValidationError(
                row=0,
                column="",
                value=len(df),
                message=f"Batch exceeds {MAX_BATCH_ROWS}-row limit ({len(df)} rows found).",
            )
        )
        return result

    if cec_db is None:
        cec_db = CECDatabase()

    # Validate unique site names
    _check_duplicate_names(df, result)

    for idx, pandas_row in df.iterrows():
        row = {k: _clean_value(v) for k, v in pandas_row.to_dict().items()}
        # Spreadsheet row = pandas index + header row(s) offset.
        # We report 1-indexed row as the user sees it: row 1 = header,
        # row 2 = possibly description, data starts at row 2 or 3 in
        # the DataFrame (0-indexed). We store the DataFrame integer
        # index and add 2 to account for the 1-indexed header row.
        display_row = int(idx) + 2  # type: ignore[arg-type]

        row_errors: list[ValidationError] = []
        row_warnings: list[ValidationError] = []

        _validate_required_fields(row, display_row, row_errors)

        analysis_type = str(row.get("analysis_type", "")).strip().lower() if row.get("analysis_type") is not None else None

        if analysis_type and analysis_type not in VALID_ANALYSIS_TYPES:
            row_errors.append(
                ValidationError(
                    row=display_row,
                    column="analysis_type",
                    value=row.get("analysis_type"),
                    message=f"Invalid analysis_type '{row.get('analysis_type')}'. Must be one of: {', '.join(sorted(VALID_ANALYSIS_TYPES))}.",
                )
            )
            # Normalize for downstream checks even though it's invalid
            analysis_type = None

        if analysis_type:
            _validate_type_required_fields(row, display_row, analysis_type, row_errors)

        _validate_field_types(row, display_row, analysis_type, row_errors, row_warnings)
        _validate_equipment(row, display_row, cec_db, row_errors)

        result.errors.extend(row_errors)
        result.warnings.extend(row_warnings)

        if not row_errors:
            result.valid_rows.append(row)

    result.is_valid = len(result.valid_rows) > 0 and len(result.errors) == 0
    return result


# ── File reading ───────────────────────────────────────────────────────


def _read_file(file_path: Path, ext: str) -> pd.DataFrame:
    """Read a CSV or Excel file into a DataFrame with normalized columns."""
    if ext == ".csv":
        df = pd.read_csv(file_path)
    else:
        df = pd.read_excel(
            file_path,
            sheet_name="Template" if _has_template_sheet(file_path) else 0,
            engine="openpyxl",
        )

    # Normalize column names: strip whitespace, lowercase
    df.columns = [str(c).strip().lower() for c in df.columns]

    # Drop columns not in the template
    known = set(TEMPLATE_COLUMNS)
    df = df[[c for c in df.columns if c in known]]

    # Skip description row if present: check if the "latitude" value
    # in the first data row is non-numeric (our template puts
    # descriptions in row 2).
    if len(df) > 0 and "latitude" in df.columns:
        first_val = df.iloc[0].get("latitude")
        if isinstance(first_val, str) and not _is_numeric(first_val):
            df = df.iloc[1:].reset_index(drop=True)

    return df


def _has_template_sheet(file_path: Path) -> bool:
    """Check if an Excel file has a sheet named 'Template'."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True)
    has_it = "Template" in wb.sheetnames
    wb.close()
    return has_it


def _is_numeric(value: Any) -> bool:
    """Check if a value can be parsed as a float."""
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False


def _clean_value(value: Any) -> Any:
    """Normalize a cell value: NaN → None, strip strings."""
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, str):
        value = value.strip()
        return None if value == "" else value
    return value


# ── Validation helpers ─────────────────────────────────────────────────


def _check_duplicate_names(df: pd.DataFrame, result: ValidationResult) -> None:
    """Add warnings for duplicate site names."""
    if "name" not in df.columns:
        return
    names = df["name"].dropna()
    dupes = names[names.duplicated(keep=False)]
    seen: set[str] = set()
    for idx, name in dupes.items():
        name_str = str(name).strip()
        if name_str in seen:
            result.warnings.append(
                ValidationError(
                    row=int(idx) + 2,
                    column="name",
                    value=name_str,
                    message=f"Duplicate site name '{name_str}'. Consider using unique names for easier identification in results.",
                )
            )
        seen.add(name_str)


def _validate_required_fields(
    row: dict, display_row: int, errors: list[ValidationError]
) -> None:
    """Check that all universally required fields are present."""
    for col in REQUIRED_FIELDS_ALL:
        val = row.get(col)
        if val is None:
            errors.append(
                ValidationError(
                    row=display_row,
                    column=col,
                    value=None,
                    message=f"Required field '{col}' is missing.",
                )
            )


def _validate_type_required_fields(
    row: dict,
    display_row: int,
    analysis_type: str,
    errors: list[ValidationError],
) -> None:
    """Check per-analysis-type required fields."""
    required = REQUIRED_FIELDS_BY_TYPE.get(analysis_type, ())
    for col in required:
        val = row.get(col)
        if val is None:
            errors.append(
                ValidationError(
                    row=display_row,
                    column=col,
                    value=None,
                    message=f"Field '{col}' is required for analysis_type '{analysis_type}'.",
                )
            )


def _validate_field_types(
    row: dict,
    display_row: int,
    analysis_type: str | None,
    errors: list[ValidationError],
    warnings: list[ValidationError],
) -> None:
    """Validate individual field types and ranges."""
    _validate_float_range(row, display_row, "latitude", -90, 90, errors)
    _validate_float_range(row, display_row, "longitude", -180, 180, errors)

    if row.get("gcr") is not None:
        _validate_float_range(row, display_row, "gcr", 0.1, 0.9, errors)
        # Warn on unusually high GCR
        gcr = _try_float(row["gcr"])
        racking = str(row.get("racking", "tracker")).strip().lower() if row.get("racking") is not None else "tracker"
        if gcr is not None:
            threshold = 0.75 if racking == "fixed" else 0.65
            if gcr > threshold:
                warnings.append(
                    ValidationError(
                        row=display_row,
                        column="gcr",
                        value=row["gcr"],
                        message=f"GCR {gcr} is unusually high for {racking} racking (typical max ~{threshold}).",
                    )
                )

    if row.get("dc_capacity_mw") is not None:
        _validate_float_gt(row, display_row, "dc_capacity_mw", 0, errors)
        dc = _try_float(row["dc_capacity_mw"])
        if dc is not None and dc > 100:
            warnings.append(
                ValidationError(
                    row=display_row,
                    column="dc_capacity_mw",
                    value=row["dc_capacity_mw"],
                    message=f"DC capacity {dc} MW is very large. Verify this is correct.",
                )
            )

    if row.get("ac_capacity_mw") is not None:
        _validate_float_gt(row, display_row, "ac_capacity_mw", 0, errors)

    # DC/AC ratio warning
    dc_val = _try_float(row.get("dc_capacity_mw"))
    ac_val = _try_float(row.get("ac_capacity_mw"))
    if dc_val is not None and ac_val is not None and ac_val > 0:
        ratio = dc_val / ac_val
        if ratio < 1.0 or ratio > 1.8:
            warnings.append(
                ValidationError(
                    row=display_row,
                    column="dc_capacity_mw",
                    value=f"DC/AC={ratio:.2f}",
                    message=f"DC/AC ratio {ratio:.2f} is outside typical range (1.0-1.8).",
                )
            )

    if row.get("buildable_acres") is not None:
        _validate_float_gt(row, display_row, "buildable_acres", 0, errors)

    if row.get("racking") is not None:
        val = str(row["racking"]).strip().lower()
        if val not in VALID_RACKING:
            errors.append(
                ValidationError(
                    row=display_row,
                    column="racking",
                    value=row["racking"],
                    message=f"Invalid racking '{row['racking']}'. Must be 'tracker' or 'fixed'.",
                )
            )

    if row.get("dispatch_mode") is not None:
        val = str(row["dispatch_mode"]).strip().lower()
        if val == "btm":
            errors.append(
                ValidationError(
                    row=display_row,
                    column="dispatch_mode",
                    value="btm",
                    message="BTM BESS is not supported in batch v1. Use dispatch_mode='ftm' or remove this row.",
                )
            )
        elif val not in VALID_DISPATCH_MODE:
            errors.append(
                ValidationError(
                    row=display_row,
                    column="dispatch_mode",
                    value=row["dispatch_mode"],
                    message=f"Invalid dispatch_mode '{row['dispatch_mode']}'. Must be 'ftm'.",
                )
            )

    if row.get("charging_mode") is not None:
        val = str(row["charging_mode"]).strip().lower()
        if val not in VALID_CHARGING_MODE:
            errors.append(
                ValidationError(
                    row=display_row,
                    column="charging_mode",
                    value=row["charging_mode"],
                    message=f"Invalid charging_mode '{row['charging_mode']}'. Must be 'solar_only' or 'solar_and_grid'.",
                )
            )

    if row.get("bifacial") is not None:
        if not _is_bool_like(row["bifacial"]):
            errors.append(
                ValidationError(
                    row=display_row,
                    column="bifacial",
                    value=row["bifacial"],
                    message=f"Invalid bifacial value '{row['bifacial']}'. Use true/false, yes/no, or 1/0.",
                )
            )

    if row.get("tilt") is not None:
        _validate_float_range(row, display_row, "tilt", 0, 90, errors)

    if row.get("azimuth") is not None:
        _validate_float_range(row, display_row, "azimuth", 0, 360, errors)

    if row.get("utilization_factor") is not None:
        _validate_float_range(row, display_row, "utilization_factor", 0, 1, errors)

    if row.get("project_lifetime_years") is not None:
        _validate_int_range(row, display_row, "project_lifetime_years", 1, 50, errors)

    # Percentage columns
    for col in PCT_COLUMNS:
        if row.get(col) is not None:
            _validate_float_range(row, display_row, col, 0, 100, errors)

    # Cost/price columns
    for col in COST_COLUMNS:
        if row.get(col) is not None:
            _validate_float_gte(row, display_row, col, 0, errors)

    # GCR sweep range columns
    for col in ("gcr_min", "gcr_max"):
        if row.get(col) is not None:
            _validate_float_range(row, display_row, col, 0.1, 0.9, errors)

    # DC:AC sweep range columns
    for col in ("dcac_min", "dcac_max"):
        if row.get(col) is not None:
            _validate_float_gt(row, display_row, col, 0, errors)


def _validate_equipment(
    row: dict,
    display_row: int,
    cec_db: CECDatabase,
    errors: list[ValidationError],
) -> None:
    """Validate module and inverter names against CEC database."""
    if row.get("module") is not None:
        module_name = str(row["module"]).strip()
        try:
            cec_db.get_module_params(module_name)
        except ModuleNotFoundError as exc:
            suggestions = exc.similar_modules[:3]
            hint = f" Did you mean: {', '.join(suggestions)}" if suggestions else ""
            errors.append(
                ValidationError(
                    row=display_row,
                    column="module",
                    value=module_name,
                    message=f"Module '{module_name}' not found in CEC database.{hint}",
                )
            )

    if row.get("inverter") is not None:
        inverter_name = str(row["inverter"]).strip()
        try:
            cec_db.get_inverter_params(inverter_name)
        except InverterNotFoundError as exc:
            suggestions = exc.similar_inverters[:3]
            hint = f" Did you mean: {', '.join(suggestions)}" if suggestions else ""
            errors.append(
                ValidationError(
                    row=display_row,
                    column="inverter",
                    value=inverter_name,
                    message=f"Inverter '{inverter_name}' not found in CEC database.{hint}",
                )
            )


# ── Type coercion helpers ──────────────────────────────────────────────


def _try_float(value: Any) -> float | None:
    """Attempt to convert value to float, returning None on failure."""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _is_bool_like(value: Any) -> bool:
    """Check if a value is bool-like (true/false/yes/no/1/0)."""
    if isinstance(value, bool):
        return True
    if isinstance(value, (int, float)):
        return value in (0, 1, 0.0, 1.0)
    if isinstance(value, str):
        return value.strip().lower() in ("true", "false", "yes", "no", "1", "0")
    return False


def _validate_float_range(
    row: dict,
    display_row: int,
    col: str,
    min_val: float,
    max_val: float,
    errors: list[ValidationError],
) -> None:
    """Validate that a column value is a float within [min_val, max_val]."""
    raw = row.get(col)
    if raw is None:
        return
    val = _try_float(raw)
    if val is None:
        errors.append(
            ValidationError(
                row=display_row,
                column=col,
                value=raw,
                message=f"'{col}' must be a number, got '{raw}'.",
            )
        )
        return
    if val < min_val or val > max_val:
        errors.append(
            ValidationError(
                row=display_row,
                column=col,
                value=raw,
                message=f"'{col}' must be between {min_val} and {max_val}, got {val}.",
            )
        )


def _validate_float_gt(
    row: dict,
    display_row: int,
    col: str,
    min_val: float,
    errors: list[ValidationError],
) -> None:
    """Validate that a column value is a float > min_val."""
    raw = row.get(col)
    if raw is None:
        return
    val = _try_float(raw)
    if val is None:
        errors.append(
            ValidationError(
                row=display_row,
                column=col,
                value=raw,
                message=f"'{col}' must be a number, got '{raw}'.",
            )
        )
        return
    if val <= min_val:
        errors.append(
            ValidationError(
                row=display_row,
                column=col,
                value=raw,
                message=f"'{col}' must be greater than {min_val}, got {val}.",
            )
        )


def _validate_float_gte(
    row: dict,
    display_row: int,
    col: str,
    min_val: float,
    errors: list[ValidationError],
) -> None:
    """Validate that a column value is a float >= min_val."""
    raw = row.get(col)
    if raw is None:
        return
    val = _try_float(raw)
    if val is None:
        errors.append(
            ValidationError(
                row=display_row,
                column=col,
                value=raw,
                message=f"'{col}' must be a number, got '{raw}'.",
            )
        )
        return
    if val < min_val:
        errors.append(
            ValidationError(
                row=display_row,
                column=col,
                value=raw,
                message=f"'{col}' must be >= {min_val}, got {val}.",
            )
        )


def _validate_int_range(
    row: dict,
    display_row: int,
    col: str,
    min_val: int,
    max_val: int,
    errors: list[ValidationError],
) -> None:
    """Validate that a column value is an integer within [min_val, max_val]."""
    raw = row.get(col)
    if raw is None:
        return
    val = _try_float(raw)
    if val is None or val != int(val):
        errors.append(
            ValidationError(
                row=display_row,
                column=col,
                value=raw,
                message=f"'{col}' must be a whole number, got '{raw}'.",
            )
        )
        return
    int_val = int(val)
    if int_val < min_val or int_val > max_val:
        errors.append(
            ValidationError(
                row=display_row,
                column=col,
                value=raw,
                message=f"'{col}' must be between {min_val} and {max_val}, got {int_val}.",
            )
        )
