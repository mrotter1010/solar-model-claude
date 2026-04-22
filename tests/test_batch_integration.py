"""End-to-end integration tests for batch processing.

These tests call REAL analysis functions (PySAM, NSRDB, USGS 3DEP, MRLC) —
no mocks. They require network access and take 1-5 minutes total.

Run with:
    pytest tests/test_batch_integration.py -v -m integration

Skip in CI/fast runs:
    pytest -m "not integration"
"""

from __future__ import annotations

import csv
import io
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.batch.executor import BatchResult, execute_batch
from src.batch.validator import validate_batch_input
from src.batch.workbook import generate_batch_workbook

# ---------------------------------------------------------------------------
# Marker — all tests in this module require network access and real backends
# ---------------------------------------------------------------------------

pytestmark = pytest.mark.integration

# ---------------------------------------------------------------------------
# Row fixtures — minimal dicts matching the template column names
# ---------------------------------------------------------------------------


@pytest.fixture()
def buildability_row() -> dict:
    """Single buildability row — Phoenix, AZ."""
    return {
        "name": "Integration Test Site",
        "latitude": 33.45,
        "longitude": -112.07,
        "analysis_type": "buildability",
    }


@pytest.fixture()
def production_row() -> dict:
    """Single production row — Denver, CO with tracker."""
    return {
        "name": "Production Test",
        "latitude": 39.74,
        "longitude": -104.99,
        "analysis_type": "production",
        "racking": "tracker",
        "gcr": 0.40,
        "dc_capacity_mw": 5.0,
        "ac_capacity_mw": 4.0,
    }


@pytest.fixture()
def optimization_row() -> dict:
    """Single optimization row — Austin, TX."""
    return {
        "name": "Optimization Test",
        "latitude": 30.27,
        "longitude": -97.74,
        "analysis_type": "optimization",
        "buildable_acres": 50.0,
        "racking": "tracker",
    }


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _write_csv(tmp_path: Path, rows: list[dict]) -> Path:
    """Write a list of row dicts to a CSV file with a union of all keys."""
    # Collect all columns in insertion order
    columns: list[str] = []
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)

    csv_path = tmp_path / "batch_input.csv"
    with open(csv_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)

    return csv_path


def _validate_and_execute(
    rows: list[dict], tmp_path: Path
) -> tuple[BatchResult, list[dict]]:
    """Write rows to CSV, validate, execute, return (result, valid_rows)."""
    csv_path = _write_csv(tmp_path, rows)
    validation = validate_batch_input(csv_path)

    assert validation.is_valid, (
        f"Validation failed unexpectedly: "
        f"{[(e.row, e.column, e.message) for e in validation.errors]}"
    )

    result = execute_batch(
        validation.valid_rows, output_base_dir=tmp_path
    )
    return result, validation.valid_rows


# ---------------------------------------------------------------------------
# Integration test cases
# ---------------------------------------------------------------------------


class TestSingleBuildabilityRow:
    """Buildability analysis for a single Phoenix, AZ site."""

    @pytest.mark.integration
    def test_single_buildability_row(
        self, buildability_row: dict, tmp_path: Path
    ) -> None:
        result, valid_rows = _validate_and_execute(
            [buildability_row], tmp_path
        )

        # Execution assertions
        assert result.total_rows == 1
        assert result.succeeded == 1
        assert result.failed == 0
        assert result.total_runtime_seconds > 0

        rr = result.row_results[0]
        assert rr.status == "success"
        assert rr.error is None
        assert rr.results is not None
        assert rr.results["buildable_acres"] > 0

        # Workbook generation and verification
        wb_path = generate_batch_workbook(result, valid_rows)
        assert wb_path.exists()

        wb = load_workbook(wb_path)
        assert "Buildability" in wb.sheetnames
        assert "Input Echo" in wb.sheetnames
        assert "Errors" not in wb.sheetnames


class TestSingleProductionRow:
    """Production (PySAM) analysis for a single Denver, CO site."""

    @pytest.mark.integration
    def test_single_production_row(
        self, production_row: dict, tmp_path: Path
    ) -> None:
        result, valid_rows = _validate_and_execute(
            [production_row], tmp_path
        )

        assert result.total_rows == 1
        assert result.succeeded == 1
        assert result.failed == 0

        rr = result.row_results[0]
        assert rr.status == "success"
        assert rr.error is None
        assert rr.results is not None
        assert rr.results["annual_energy_mwh"] > 0

        cf = rr.results["capacity_factor_ac_pct"]
        assert cf is not None
        # Denver tracker at GCR 0.40 should be ~20-30% (0-100 scale)
        assert 10 <= cf <= 40, f"Unexpected capacity factor: {cf}"

        # Workbook
        wb_path = generate_batch_workbook(result, valid_rows)
        wb = load_workbook(wb_path)
        assert "Production" in wb.sheetnames
        assert "Errors" not in wb.sheetnames


class TestSingleOptimizationRow:
    """Solar layout optimization for a single Austin, TX site."""

    @pytest.mark.integration
    def test_single_optimization_row(
        self, optimization_row: dict, tmp_path: Path
    ) -> None:
        result, valid_rows = _validate_and_execute(
            [optimization_row], tmp_path
        )

        assert result.total_rows == 1
        assert result.succeeded == 1
        assert result.failed == 0

        rr = result.row_results[0]
        assert rr.status == "success"
        assert rr.error is None
        assert rr.results is not None
        assert rr.results["winner_gcr"] > 0
        assert rr.results["winner_mw_dc"] > 0

        # Workbook
        wb_path = generate_batch_workbook(result, valid_rows)
        wb = load_workbook(wb_path)
        assert "Optimization" in wb.sheetnames
        assert "Errors" not in wb.sheetnames


class TestMixedBatch:
    """Three rows: buildability + production + optimization."""

    @pytest.mark.integration
    def test_mixed_batch(
        self,
        buildability_row: dict,
        production_row: dict,
        optimization_row: dict,
        tmp_path: Path,
    ) -> None:
        rows = [buildability_row, production_row, optimization_row]
        result, valid_rows = _validate_and_execute(rows, tmp_path)

        assert result.total_rows == 3
        assert result.succeeded == 3
        assert result.failed == 0

        # Verify each row succeeded
        for rr in result.row_results:
            assert rr.status == "success", (
                f"{rr.name} ({rr.analysis_type}) failed: {rr.error}"
            )

        # Workbook should have all 3 analysis tabs + Input Echo, no Errors
        wb_path = generate_batch_workbook(result, valid_rows)
        wb = load_workbook(wb_path)
        assert "Buildability" in wb.sheetnames
        assert "Production" in wb.sheetnames
        assert "Optimization" in wb.sheetnames
        assert "Input Echo" in wb.sheetnames
        assert "Errors" not in wb.sheetnames


class TestFailForward:
    """One valid buildability row + one that passes validation but fails at runtime."""

    @pytest.mark.integration
    def test_fail_forward(
        self, buildability_row: dict, tmp_path: Path
    ) -> None:
        # Row that passes validation (valid lat/lon range, valid type)
        # but will fail at runtime — NSRDB has no data at (0, 0) mid-Atlantic,
        # and buildability analysis will fail on elevation/land cover retrieval.
        bad_row = {
            "name": "Mid-Atlantic Fail",
            "latitude": 0.0,
            "longitude": 0.0,
            "analysis_type": "buildability",
        }

        csv_path = _write_csv(tmp_path, [buildability_row, bad_row])
        validation = validate_batch_input(csv_path)

        assert validation.is_valid, (
            f"Validation should pass: "
            f"{[(e.row, e.column, e.message) for e in validation.errors]}"
        )

        result = execute_batch(
            validation.valid_rows, output_base_dir=tmp_path
        )

        assert result.total_rows == 2
        assert result.succeeded == 1
        assert result.failed == 1

        # First row (Phoenix) should succeed
        assert result.row_results[0].status == "success"
        assert result.row_results[0].results is not None
        assert result.row_results[0].results["buildable_acres"] > 0

        # Second row (0,0) should fail
        assert result.row_results[1].status == "failed"
        assert result.row_results[1].error is not None

        # Workbook should have Errors tab for the failed row
        wb_path = generate_batch_workbook(
            result, validation.valid_rows
        )
        wb = load_workbook(wb_path)
        assert "Errors" in wb.sheetnames
        assert "Buildability" in wb.sheetnames


class TestFullApiEndpoint:
    """End-to-end test through FastAPI endpoints."""

    @pytest.mark.integration
    def test_full_api_endpoint(self, tmp_path: Path) -> None:
        from fastapi.testclient import TestClient

        from src.api.app import create_app

        app = create_app()
        client = TestClient(app)

        # Build a 2-row buildability CSV (fastest analysis type)
        csv_content = (
            "name,latitude,longitude,analysis_type\n"
            "API Site A,33.45,-112.07,buildability\n"
            "API Site B,32.22,-110.97,buildability\n"
        )

        response = client.post(
            "/analyses/batch",
            files={
                "file": (
                    "batch_input.csv",
                    io.BytesIO(csv_content.encode()),
                    "text/csv",
                )
            },
        )

        assert response.status_code == 200, (
            f"Batch POST failed: {response.status_code} {response.text}"
        )

        data = response.json()
        assert "run_id" in data
        assert data["total_rows"] == 2
        assert data["succeeded"] == 2
        assert data["failed"] == 0
        assert "files" in data
        assert "workbook_url" in data["files"]

        # Download the workbook
        workbook_url = data["files"]["workbook_url"]
        dl_response = client.get(workbook_url)
        assert dl_response.status_code == 200
        assert "spreadsheetml" in dl_response.headers.get(
            "content-type", ""
        )


class TestProgressCallback:
    """Verify the progress_callback is called for each row."""

    @pytest.mark.integration
    def test_progress_callback(self, tmp_path: Path) -> None:
        # Three minimal buildability rows (fastest analysis type)
        rows = [
            {
                "name": "Progress A",
                "latitude": 33.45,
                "longitude": -112.07,
                "analysis_type": "buildability",
            },
            {
                "name": "Progress B",
                "latitude": 32.22,
                "longitude": -110.97,
                "analysis_type": "buildability",
            },
            {
                "name": "Progress C",
                "latitude": 34.05,
                "longitude": -111.09,
                "analysis_type": "buildability",
            },
        ]

        csv_path = _write_csv(tmp_path, rows)
        validation = validate_batch_input(csv_path)
        assert validation.is_valid

        # Track callback calls
        calls: list[tuple[int, int, str, str]] = []

        def callback(
            current: int, total: int, name: str, status: str
        ) -> None:
            calls.append((current, total, name, status))

        result = execute_batch(
            validation.valid_rows,
            output_base_dir=tmp_path,
            progress_callback=callback,
        )

        # Should be called once per row
        assert len(calls) == 3

        # Verify (current, total) progression
        assert calls[0][0] == 1
        assert calls[0][1] == 3
        assert calls[1][0] == 2
        assert calls[1][1] == 3
        assert calls[2][0] == 3
        assert calls[2][1] == 3

        # Names match input
        assert calls[0][2] == "Progress A"
        assert calls[1][2] == "Progress B"
        assert calls[2][2] == "Progress C"

        # All should succeed (status is either "success" or "failed")
        for current, total, name, status in calls:
            assert status == "success", (
                f"Row {name} status was {status}, expected success"
            )

        # Overall result matches
        assert result.succeeded == 3
        assert result.failed == 0
