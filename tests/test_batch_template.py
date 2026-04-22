"""Tests for batch processing template generation."""

from __future__ import annotations

import csv
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.batch.template import (
    EXAMPLE_ROWS,
    TEMPLATE_COLUMNS,
    generate_batch_template,
)


@pytest.fixture
def tmp_output_dir(tmp_path: Path) -> Path:
    """Provide a temporary directory for generated templates."""
    return tmp_path


# ── Excel template tests ───────────────────────────────────────────────


class TestExcelTemplate:
    """Tests for .xlsx template generation."""

    def test_creates_valid_xlsx_file(self, tmp_output_dir: Path) -> None:
        """generate_batch_template creates a readable .xlsx file."""
        result = generate_batch_template(tmp_output_dir, format="xlsx")

        assert result.exists()
        assert result.suffix == ".xlsx"
        # Verify openpyxl can read it back
        wb = load_workbook(result)
        assert wb.sheetnames[0] == "Template"
        wb.close()

    def test_template_sheet_has_correct_headers(self, tmp_output_dir: Path) -> None:
        """Row 1 of the Template sheet matches TEMPLATE_COLUMNS exactly."""
        result = generate_batch_template(tmp_output_dir, format="xlsx")
        wb = load_workbook(result)
        ws = wb["Template"]

        headers = [ws.cell(row=1, column=i).value for i in range(1, len(TEMPLATE_COLUMNS) + 1)]
        assert headers == TEMPLATE_COLUMNS
        wb.close()

    def test_has_four_example_rows(self, tmp_output_dir: Path) -> None:
        """Rows 3-6 contain four example rows with data."""
        result = generate_batch_template(tmp_output_dir, format="xlsx")
        wb = load_workbook(result)
        ws = wb["Template"]

        # Row 1 = headers, Row 2 = descriptions, Rows 3-6 = examples
        example_names = []
        for row_num in range(3, 7):
            name = ws.cell(row=row_num, column=1).value  # 'name' is column 1
            if name is not None:
                example_names.append(name)

        assert len(example_names) == len(EXAMPLE_ROWS)
        assert example_names == [row["name"] for row in EXAMPLE_ROWS]
        wb.close()

    def test_example_analysis_types(self, tmp_output_dir: Path) -> None:
        """Each example row has the expected analysis_type."""
        result = generate_batch_template(tmp_output_dir, format="xlsx")
        wb = load_workbook(result)
        ws = wb["Template"]

        # analysis_type is column 4
        analysis_types = [ws.cell(row=r, column=4).value for r in range(3, 7)]
        assert analysis_types == [
            "buildability",
            "production",
            "optimization",
            "optimization_bess",
        ]
        wb.close()

    def test_instructions_sheet_exists(self, tmp_output_dir: Path) -> None:
        """The workbook has an Instructions sheet with content."""
        result = generate_batch_template(tmp_output_dir, format="xlsx")
        wb = load_workbook(result)

        assert "Instructions" in wb.sheetnames
        ws_instr = wb["Instructions"]
        # First cell should have the title text
        assert "BATCH TEMPLATE INSTRUCTIONS" in (ws_instr.cell(row=1, column=1).value or "")
        wb.close()

    def test_row_count(self, tmp_output_dir: Path) -> None:
        """Template sheet has 6 rows: 1 header + 1 description + 4 examples."""
        result = generate_batch_template(tmp_output_dir, format="xlsx")
        wb = load_workbook(result)
        ws = wb["Template"]

        assert ws.max_row == 6
        wb.close()

    def test_description_row_present(self, tmp_output_dir: Path) -> None:
        """Row 2 has column descriptions (not empty)."""
        result = generate_batch_template(tmp_output_dir, format="xlsx")
        wb = load_workbook(result)
        ws = wb["Template"]

        # Check first few description cells are populated
        assert ws.cell(row=2, column=1).value is not None  # name description
        assert ws.cell(row=2, column=4).value is not None  # analysis_type description
        wb.close()

    def test_header_row_frozen(self, tmp_output_dir: Path) -> None:
        """The Template sheet has frozen panes at row 1."""
        result = generate_batch_template(tmp_output_dir, format="xlsx")
        wb = load_workbook(result)
        ws = wb["Template"]

        assert ws.freeze_panes == "A2"
        wb.close()

    def test_explicit_file_path(self, tmp_output_dir: Path) -> None:
        """Accepts an explicit file path instead of a directory."""
        explicit = tmp_output_dir / "my_template.xlsx"
        result = generate_batch_template(explicit, format="xlsx")

        assert result == explicit
        assert result.exists()


# ── CSV template tests ─────────────────────────────────────────────────


class TestCSVTemplate:
    """Tests for .csv template generation."""

    def test_creates_valid_csv_file(self, tmp_output_dir: Path) -> None:
        """generate_batch_template creates a readable .csv file."""
        result = generate_batch_template(tmp_output_dir, format="csv")

        assert result.exists()
        assert result.suffix == ".csv"

    def test_csv_has_correct_headers(self, tmp_output_dir: Path) -> None:
        """CSV contains the correct column headers."""
        result = generate_batch_template(tmp_output_dir, format="csv")

        with open(result) as f:
            reader = csv.reader(f)
            headers = next(reader)

        assert headers == TEMPLATE_COLUMNS

    def test_csv_header_only(self, tmp_output_dir: Path) -> None:
        """CSV has only the header row (no descriptions or examples)."""
        result = generate_batch_template(tmp_output_dir, format="csv")

        with open(result) as f:
            reader = csv.reader(f)
            rows = list(reader)

        assert len(rows) == 1


# ── Error handling ─────────────────────────────────────────────────────


class TestTemplateErrors:
    """Tests for error cases."""

    def test_invalid_format_raises(self, tmp_output_dir: Path) -> None:
        """Unsupported format raises ValueError."""
        with pytest.raises(ValueError, match="Unsupported format"):
            generate_batch_template(tmp_output_dir, format="pdf")
