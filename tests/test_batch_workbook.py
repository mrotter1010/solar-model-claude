"""Tests for the batch results Excel workbook generator.

All tests use mock BatchResult / RowResult objects — no real analyses run.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from src.batch.executor import BatchResult, RowResult
from src.batch.workbook import generate_batch_workbook


# ── Helpers ────────────────────────────────────────────────────────────


def _make_row_result(
    row_index: int = 1,
    name: str = "Site A",
    analysis_type: str = "buildability",
    status: str = "success",
    error: str | None = None,
    runtime_seconds: float = 1.5,
    results: dict | None = None,
) -> RowResult:
    return RowResult(
        row_index=row_index,
        name=name,
        analysis_type=analysis_type,
        status=status,
        error=error,
        runtime_seconds=runtime_seconds,
        results=results,
    )


def _make_batch_result(
    row_results: list[RowResult],
    output_dir: Path,
    run_id: str = "abc123def456",
) -> BatchResult:
    succeeded = sum(1 for r in row_results if r.status == "success")
    failed = sum(1 for r in row_results if r.status == "failed")
    return BatchResult(
        run_id=run_id,
        total_rows=len(row_results),
        succeeded=succeeded,
        failed=failed,
        total_runtime_seconds=5.0,
        row_results=row_results,
        output_dir=output_dir,
    )


def _buildability_results() -> dict:
    return {
        "buildable_acres": 350.0,
        "buildable_pct": 70.0,
        "total_acres": 500.0,
        "dominant_land_cover": "Grassland/Herbaceous",
        "mean_slope": 3.5,
        "tracker_suitable_pct": 85.0,
    }


def _production_results() -> dict:
    return {
        "annual_energy_mwh": 10500.0,
        "capacity_factor_ac_pct": 25.4,
        "specific_yield_kwh_per_kw": 1680.0,
        "performance_ratio_pct": 82.0,
        "dc_capacity_mw": 5.0,
        "ac_capacity_mw": 4.0,
        "gcr": 0.40,
        "dcac_ratio": 1.25,
    }


def _optimization_results() -> dict:
    return {
        "winner_type": "production",
        "winner_gcr": 0.45,
        "winner_dcac": 1.30,
        "winner_mw_dc": 12.5,
        "winner_mw_ac": 9.6,
        "winner_annual_mwh": 21000.0,
        "winner_cf_pct": 24.9,
        "winner_lcoe_per_mwh": None,
        "winner_npv": None,
        "buildable_acres": 100.0,
    }


def _optimization_bess_results() -> dict:
    return {
        "winner_type": "npv",
        "winner_gcr": 0.40,
        "winner_dcac": 1.25,
        "winner_mw_dc": 10.0,
        "winner_mw_ac": 8.0,
        "winner_annual_mwh": 18000.0,
        "winner_cf_pct": 25.7,
        "winner_lcoe_per_mwh": 35.0,
        "winner_npv": 2500000.0,
        "buildable_acres": 80.0,
        "bess_adds_value": True,
        "bess_power_mw": 2.0,
        "bess_duration_hr": 4.0,
        "bess_npv": 700000.0,
        "combined_npv": 3200000.0,
    }


def _buildability_input_row() -> dict:
    return {
        "name": "Site A",
        "latitude": 33.45,
        "longitude": -112.07,
        "analysis_type": "buildability",
    }


def _production_input_row() -> dict:
    return {
        "name": "Site B",
        "latitude": 39.74,
        "longitude": -104.99,
        "analysis_type": "production",
        "racking": "tracker",
        "gcr": 0.40,
        "dc_capacity_mw": 5.0,
        "ac_capacity_mw": 4.0,
    }


def _optimization_input_row() -> dict:
    return {
        "name": "Site C",
        "latitude": 30.27,
        "longitude": -97.74,
        "analysis_type": "optimization",
        "buildable_acres": 100.0,
    }


def _bess_input_row() -> dict:
    return {
        "name": "Site D",
        "latitude": 39.95,
        "longitude": -75.16,
        "analysis_type": "optimization_bess",
        "buildable_acres": 80.0,
        "dispatch_mode": "ftm",
    }


# ── Tests: buildability only ──────────────────────────────────────────


class TestBuildabilityOnly:
    """Workbook with buildability results only."""

    def test_creates_buildability_and_input_echo_tabs(self, tmp_path: Path) -> None:
        rr = _make_row_result(results=_buildability_results())
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_buildability_input_row()])

        wb = load_workbook(path)
        assert "Buildability" in wb.sheetnames
        assert "Input Echo" in wb.sheetnames
        assert "Errors" not in wb.sheetnames
        assert len(wb.sheetnames) == 2

    def test_buildability_headers(self, tmp_path: Path) -> None:
        rr = _make_row_result(results=_buildability_results())
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_buildability_input_row()])

        wb = load_workbook(path)
        ws = wb["Buildability"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Site Name" in headers
        assert "Latitude" in headers
        assert "Buildable Acres" in headers
        assert "Tracker Suitable %" in headers

    def test_buildability_data_values(self, tmp_path: Path) -> None:
        rr = _make_row_result(results=_buildability_results())
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_buildability_input_row()])

        wb = load_workbook(path)
        ws = wb["Buildability"]
        # Row 2 is the first data row
        # Find column indices by header
        col_map = {
            ws.cell(row=1, column=c).value: c
            for c in range(1, ws.max_column + 1)
        }
        assert ws.cell(row=2, column=col_map["Site Name"]).value == "Site A"
        assert ws.cell(row=2, column=col_map["Buildable Acres"]).value == 350.0
        assert ws.cell(row=2, column=col_map["Buildable %"]).value == 70.0
        assert ws.cell(row=2, column=col_map["Total Acres"]).value == 500.0
        assert ws.cell(row=2, column=col_map["Status"]).value == "success"


# ── Tests: production only ─────────────────────────────────────────────


class TestProductionOnly:
    """Workbook with production results only."""

    def test_creates_production_tab(self, tmp_path: Path) -> None:
        rr = _make_row_result(
            name="Site B",
            analysis_type="production",
            results=_production_results(),
        )
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_production_input_row()])

        wb = load_workbook(path)
        assert "Production" in wb.sheetnames
        assert "Buildability" not in wb.sheetnames

    def test_production_data_values(self, tmp_path: Path) -> None:
        rr = _make_row_result(
            name="Site B",
            analysis_type="production",
            results=_production_results(),
        )
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_production_input_row()])

        wb = load_workbook(path)
        ws = wb["Production"]
        col_map = {
            ws.cell(row=1, column=c).value: c
            for c in range(1, ws.max_column + 1)
        }
        assert ws.cell(row=2, column=col_map["Annual Energy (MWh)"]).value == 10500.0
        assert ws.cell(row=2, column=col_map["DC Capacity (MW)"]).value == 5.0
        assert ws.cell(row=2, column=col_map["GCR"]).value == 0.40
        assert ws.cell(row=2, column=col_map["DC:AC Ratio"]).value == 1.25


# ── Tests: mixed types ─────────────────────────────────────────────────


class TestMixedTypes:
    """Workbook with multiple analysis types."""

    def test_creates_multiple_analysis_tabs(self, tmp_path: Path) -> None:
        rows = [
            _make_row_result(
                row_index=1,
                name="Site A",
                analysis_type="buildability",
                results=_buildability_results(),
            ),
            _make_row_result(
                row_index=2,
                name="Site B",
                analysis_type="production",
                results=_production_results(),
            ),
            _make_row_result(
                row_index=3,
                name="Site C",
                analysis_type="optimization",
                results=_optimization_results(),
            ),
        ]
        br = _make_batch_result(rows, tmp_path)
        input_rows = [
            _buildability_input_row(),
            _production_input_row(),
            _optimization_input_row(),
        ]
        path = generate_batch_workbook(br, input_rows)

        wb = load_workbook(path)
        assert "Buildability" in wb.sheetnames
        assert "Production" in wb.sheetnames
        assert "Optimization" in wb.sheetnames
        assert "Optimization BESS" not in wb.sheetnames
        assert "Input Echo" in wb.sheetnames

    def test_all_four_types(self, tmp_path: Path) -> None:
        rows = [
            _make_row_result(
                row_index=1,
                name="Site A",
                analysis_type="buildability",
                results=_buildability_results(),
            ),
            _make_row_result(
                row_index=2,
                name="Site B",
                analysis_type="production",
                results=_production_results(),
            ),
            _make_row_result(
                row_index=3,
                name="Site C",
                analysis_type="optimization",
                results=_optimization_results(),
            ),
            _make_row_result(
                row_index=4,
                name="Site D",
                analysis_type="optimization_bess",
                results=_optimization_bess_results(),
            ),
        ]
        br = _make_batch_result(rows, tmp_path)
        input_rows = [
            _buildability_input_row(),
            _production_input_row(),
            _optimization_input_row(),
            _bess_input_row(),
        ]
        path = generate_batch_workbook(br, input_rows)

        wb = load_workbook(path)
        assert len(wb.sheetnames) == 5  # 4 analysis + Input Echo
        assert "Optimization BESS" in wb.sheetnames

    def test_each_type_tab_has_correct_row_count(self, tmp_path: Path) -> None:
        """Each analysis tab contains only rows of that type."""
        rows = [
            _make_row_result(row_index=1, name="A1", analysis_type="buildability", results=_buildability_results()),
            _make_row_result(row_index=2, name="A2", analysis_type="buildability", results=_buildability_results()),
            _make_row_result(row_index=3, name="B1", analysis_type="production", results=_production_results()),
        ]
        br = _make_batch_result(rows, tmp_path)
        input_rows = [
            {**_buildability_input_row(), "name": "A1"},
            {**_buildability_input_row(), "name": "A2"},
            {**_production_input_row(), "name": "B1"},
        ]
        path = generate_batch_workbook(br, input_rows)

        wb = load_workbook(path)
        # Buildability: header + 2 data rows
        assert wb["Buildability"].max_row == 3
        # Production: header + 1 data row
        assert wb["Production"].max_row == 2


# ── Tests: failed rows ─────────────────────────────────────────────────


class TestFailedRows:
    """Workbook behavior when rows have failed."""

    def test_errors_tab_created_when_failures(self, tmp_path: Path) -> None:
        rows = [
            _make_row_result(
                row_index=1,
                name="Good Site",
                analysis_type="buildability",
                results=_buildability_results(),
            ),
            _make_row_result(
                row_index=2,
                name="Bad Site",
                analysis_type="buildability",
                status="failed",
                error="NLCD fetch failed",
                results=None,
            ),
        ]
        br = _make_batch_result(rows, tmp_path)
        input_rows = [
            {**_buildability_input_row(), "name": "Good Site"},
            {**_buildability_input_row(), "name": "Bad Site"},
        ]
        path = generate_batch_workbook(br, input_rows)

        wb = load_workbook(path)
        assert "Errors" in wb.sheetnames

    def test_errors_tab_content(self, tmp_path: Path) -> None:
        rows = [
            _make_row_result(
                row_index=3,
                name="Bad Site",
                analysis_type="production",
                status="failed",
                error="PySAM crashed",
            ),
        ]
        br = _make_batch_result(rows, tmp_path)
        path = generate_batch_workbook(br, [_production_input_row()])

        wb = load_workbook(path)
        ws = wb["Errors"]
        col_map = {
            ws.cell(row=1, column=c).value: c
            for c in range(1, ws.max_column + 1)
        }
        assert ws.cell(row=2, column=col_map["row_number"]).value == 3
        assert ws.cell(row=2, column=col_map["name"]).value == "Bad Site"
        assert ws.cell(row=2, column=col_map["error_message"]).value == "PySAM crashed"

    def test_failed_rows_appear_in_analysis_tab(self, tmp_path: Path) -> None:
        """Failed rows show up in analysis tab with status=failed, blank results."""
        rows = [
            _make_row_result(
                row_index=1,
                name="Good",
                analysis_type="buildability",
                results=_buildability_results(),
            ),
            _make_row_result(
                row_index=2,
                name="Bad",
                analysis_type="buildability",
                status="failed",
                error="Timeout",
            ),
        ]
        br = _make_batch_result(rows, tmp_path)
        input_rows = [
            {**_buildability_input_row(), "name": "Good"},
            {**_buildability_input_row(), "name": "Bad"},
        ]
        path = generate_batch_workbook(br, input_rows)

        wb = load_workbook(path)
        ws = wb["Buildability"]
        col_map = {
            ws.cell(row=1, column=c).value: c
            for c in range(1, ws.max_column + 1)
        }
        # Row 2 = Good, Row 3 = Bad
        assert ws.cell(row=3, column=col_map["Status"]).value == "failed"
        assert ws.cell(row=3, column=col_map["Error"]).value == "Timeout"
        # Result columns should be None/empty for failed row
        assert ws.cell(row=3, column=col_map["Buildable Acres"]).value is None

    def test_all_rows_failed(self, tmp_path: Path) -> None:
        """Analysis tabs are still created even if all rows of that type failed."""
        rows = [
            _make_row_result(
                row_index=1,
                name="F1",
                analysis_type="production",
                status="failed",
                error="Error 1",
            ),
            _make_row_result(
                row_index=2,
                name="F2",
                analysis_type="production",
                status="failed",
                error="Error 2",
            ),
        ]
        br = _make_batch_result(rows, tmp_path)
        input_rows = [
            {**_production_input_row(), "name": "F1"},
            {**_production_input_row(), "name": "F2"},
        ]
        path = generate_batch_workbook(br, input_rows)

        wb = load_workbook(path)
        assert "Production" in wb.sheetnames
        assert "Errors" in wb.sheetnames
        ws = wb["Production"]
        # Header + 2 failed data rows
        assert ws.max_row == 3


# ── Tests: formatting ──���──────────────────────────────────���────────────


class TestFormatting:
    """Verify Excel formatting details."""

    def test_frozen_rows(self, tmp_path: Path) -> None:
        rr = _make_row_result(results=_buildability_results())
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_buildability_input_row()])

        wb = load_workbook(path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            assert ws.freeze_panes == "A2", f"Freeze panes not set on {sheet_name}"

    def test_auto_filter(self, tmp_path: Path) -> None:
        rr = _make_row_result(results=_buildability_results())
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_buildability_input_row()])

        wb = load_workbook(path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            assert ws.auto_filter.ref is not None, f"Auto-filter not set on {sheet_name}"

    def test_header_bold(self, tmp_path: Path) -> None:
        rr = _make_row_result(results=_buildability_results())
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_buildability_input_row()])

        wb = load_workbook(path)
        ws = wb["Buildability"]
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            assert cell.font.bold, f"Header col {col} not bold"

    def test_header_fill_blue(self, tmp_path: Path) -> None:
        rr = _make_row_result(results=_buildability_results())
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_buildability_input_row()])

        wb = load_workbook(path)
        ws = wb["Buildability"]
        cell = ws.cell(row=1, column=1)
        # openpyxl normalizes colors to uppercase with 00 alpha prefix
        assert cell.fill.start_color.rgb in ("00D6EAF8", "D6EAF8")

    def test_errors_header_fill_red(self, tmp_path: Path) -> None:
        rr = _make_row_result(
            status="failed", error="Boom", analysis_type="buildability"
        )
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_buildability_input_row()])

        wb = load_workbook(path)
        ws = wb["Errors"]
        cell = ws.cell(row=1, column=1)
        assert cell.fill.start_color.rgb in ("00FADBD8", "FADBD8")

    def test_alternating_row_shading(self, tmp_path: Path) -> None:
        """Even-numbered rows (0-indexed data row 2, 4, ...) get light gray fill."""
        rows = [
            _make_row_result(row_index=1, name="A", analysis_type="buildability", results=_buildability_results()),
            _make_row_result(row_index=2, name="B", analysis_type="buildability", results=_buildability_results()),
            _make_row_result(row_index=3, name="C", analysis_type="buildability", results=_buildability_results()),
        ]
        br = _make_batch_result(rows, tmp_path)
        input_rows = [
            {**_buildability_input_row(), "name": "A"},
            {**_buildability_input_row(), "name": "B"},
            {**_buildability_input_row(), "name": "C"},
        ]
        path = generate_batch_workbook(br, input_rows)

        wb = load_workbook(path)
        ws = wb["Buildability"]
        # Row 2 = even → shaded; Row 3 = odd → no shading; Row 4 = even → shaded
        cell_r2 = ws.cell(row=2, column=1)
        cell_r3 = ws.cell(row=3, column=1)
        cell_r4 = ws.cell(row=4, column=1)
        assert cell_r2.fill.start_color.rgb in ("00F8F9FA", "F8F9FA")
        assert cell_r3.fill.start_color.rgb not in ("00F8F9FA", "F8F9FA")
        assert cell_r4.fill.start_color.rgb in ("00F8F9FA", "F8F9FA")

    def test_tab_colors(self, tmp_path: Path) -> None:
        rows = [
            _make_row_result(row_index=1, name="A", analysis_type="buildability", results=_buildability_results()),
            _make_row_result(row_index=2, name="F", analysis_type="production", status="failed", error="err"),
        ]
        br = _make_batch_result(rows, tmp_path)
        input_rows = [_buildability_input_row(), _production_input_row()]
        path = generate_batch_workbook(br, input_rows)

        wb = load_workbook(path)
        # Analysis tabs: blue
        assert wb["Buildability"].sheet_properties.tabColor.rgb in ("003498DB", "3498DB")
        # Input Echo: gray
        assert wb["Input Echo"].sheet_properties.tabColor.rgb in ("0095A5A6", "95A5A6")
        # Errors: red
        assert wb["Errors"].sheet_properties.tabColor.rgb in ("00E74C3C", "E74C3C")


# ── Tests: number formatting ───��──────────────────────────────────────


class TestNumberFormatting:
    """Spot-check number formats on specific cells."""

    def test_mwh_format(self, tmp_path: Path) -> None:
        rr = _make_row_result(
            name="Site B",
            analysis_type="production",
            results=_production_results(),
        )
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_production_input_row()])

        wb = load_workbook(path)
        ws = wb["Production"]
        col_map = {
            ws.cell(row=1, column=c).value: c
            for c in range(1, ws.max_column + 1)
        }
        cell = ws.cell(row=2, column=col_map["Annual Energy (MWh)"])
        assert cell.number_format == '#,##0'

    def test_pct_format(self, tmp_path: Path) -> None:
        rr = _make_row_result(results=_buildability_results())
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_buildability_input_row()])

        wb = load_workbook(path)
        ws = wb["Buildability"]
        col_map = {
            ws.cell(row=1, column=c).value: c
            for c in range(1, ws.max_column + 1)
        }
        cell = ws.cell(row=2, column=col_map["Buildable %"])
        assert cell.number_format == '0.0"%"'

    def test_gcr_format(self, tmp_path: Path) -> None:
        rr = _make_row_result(
            name="Site B",
            analysis_type="production",
            results=_production_results(),
        )
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_production_input_row()])

        wb = load_workbook(path)
        ws = wb["Production"]
        col_map = {
            ws.cell(row=1, column=c).value: c
            for c in range(1, ws.max_column + 1)
        }
        cell = ws.cell(row=2, column=col_map["GCR"])
        assert cell.number_format == '0.00'

    def test_dollar_format(self, tmp_path: Path) -> None:
        rr = _make_row_result(
            name="Site D",
            analysis_type="optimization_bess",
            results=_optimization_bess_results(),
        )
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_bess_input_row()])

        wb = load_workbook(path)
        ws = wb["Optimization BESS"]
        col_map = {
            ws.cell(row=1, column=c).value: c
            for c in range(1, ws.max_column + 1)
        }
        cell = ws.cell(row=2, column=col_map["Winner NPV ($)"])
        assert cell.number_format == '#,##0.00'

    def test_acres_format(self, tmp_path: Path) -> None:
        rr = _make_row_result(results=_buildability_results())
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_buildability_input_row()])

        wb = load_workbook(path)
        ws = wb["Buildability"]
        col_map = {
            ws.cell(row=1, column=c).value: c
            for c in range(1, ws.max_column + 1)
        }
        cell = ws.cell(row=2, column=col_map["Buildable Acres"])
        assert cell.number_format == '#,##0.0'

    def test_no_format_on_none_value(self, tmp_path: Path) -> None:
        """Cells with None values should not get number formatting applied."""
        rr = _make_row_result(
            name="Site C",
            analysis_type="optimization",
            results=_optimization_results(),  # winner_npv is None
        )
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_optimization_input_row()])

        wb = load_workbook(path)
        ws = wb["Optimization"]
        col_map = {
            ws.cell(row=1, column=c).value: c
            for c in range(1, ws.max_column + 1)
        }
        cell = ws.cell(row=2, column=col_map["Winner NPV ($)"])
        assert cell.value is None
        # Format string should still be General (not applied)
        assert cell.number_format == "General"


# ── Tests: Input Echo ────────────────────────────────────���─────────────


class TestInputEcho:
    """Verify Input Echo tab content."""

    def test_contains_all_input_columns(self, tmp_path: Path) -> None:
        input_row = _production_input_row()
        rr = _make_row_result(
            name="Site B", analysis_type="production", results=_production_results()
        )
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [input_row])

        wb = load_workbook(path)
        ws = wb["Input Echo"]
        headers = [
            ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
        ]
        for key in input_row:
            assert key in headers, f"Input column '{key}' missing from Input Echo"

    def test_raw_values_preserved(self, tmp_path: Path) -> None:
        input_row = _production_input_row()
        rr = _make_row_result(
            name="Site B", analysis_type="production", results=_production_results()
        )
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [input_row])

        wb = load_workbook(path)
        ws = wb["Input Echo"]
        col_map = {
            ws.cell(row=1, column=c).value: c
            for c in range(1, ws.max_column + 1)
        }
        assert ws.cell(row=2, column=col_map["name"]).value == "Site B"
        assert ws.cell(row=2, column=col_map["latitude"]).value == 39.74
        assert ws.cell(row=2, column=col_map["gcr"]).value == 0.40

    def test_multiple_rows(self, tmp_path: Path) -> None:
        input_rows = [
            _buildability_input_row(),
            _production_input_row(),
        ]
        rows = [
            _make_row_result(row_index=1, name="Site A", analysis_type="buildability", results=_buildability_results()),
            _make_row_result(row_index=2, name="Site B", analysis_type="production", results=_production_results()),
        ]
        br = _make_batch_result(rows, tmp_path)
        path = generate_batch_workbook(br, input_rows)

        wb = load_workbook(path)
        ws = wb["Input Echo"]
        # Header + 2 data rows
        assert ws.max_row == 3

    def test_empty_input_rows(self, tmp_path: Path) -> None:
        """Gracefully handle empty input_rows list."""
        br = _make_batch_result([], tmp_path)
        path = generate_batch_workbook(br, [])

        wb = load_workbook(path)
        assert "Input Echo" in wb.sheetnames


# ── Tests: output path ─────────────────────────────────────────────────


class TestOutputPath:
    """Verify output path behavior."""

    def test_default_path(self, tmp_path: Path) -> None:
        rr = _make_row_result(results=_buildability_results())
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_buildability_input_row()])

        assert path == tmp_path / "batch_results.xlsx"
        assert path.exists()

    def test_custom_path(self, tmp_path: Path) -> None:
        rr = _make_row_result(results=_buildability_results())
        br = _make_batch_result([rr], tmp_path)
        custom = tmp_path / "custom" / "my_report.xlsx"
        path = generate_batch_workbook(br, [_buildability_input_row()], output_path=custom)

        assert path == custom
        assert path.exists()

    def test_returns_path(self, tmp_path: Path) -> None:
        rr = _make_row_result(results=_buildability_results())
        br = _make_batch_result([rr], tmp_path)
        result = generate_batch_workbook(br, [_buildability_input_row()])
        assert isinstance(result, Path)


# ── Tests: tab ordering ────────────────────────────────────────────────


class TestTabOrdering:
    """Verify tabs appear in the correct order."""

    def test_analysis_tabs_in_canonical_order(self, tmp_path: Path) -> None:
        """buildability, production, optimization, optimization_bess, Input Echo."""
        rows = [
            _make_row_result(row_index=1, name="D", analysis_type="optimization_bess", results=_optimization_bess_results()),
            _make_row_result(row_index=2, name="A", analysis_type="buildability", results=_buildability_results()),
            _make_row_result(row_index=3, name="C", analysis_type="optimization", results=_optimization_results()),
            _make_row_result(row_index=4, name="B", analysis_type="production", results=_production_results()),
        ]
        br = _make_batch_result(rows, tmp_path)
        input_rows = [
            _bess_input_row(),
            _buildability_input_row(),
            _optimization_input_row(),
            _production_input_row(),
        ]
        path = generate_batch_workbook(br, input_rows)

        wb = load_workbook(path)
        assert wb.sheetnames == [
            "Buildability",
            "Production",
            "Optimization",
            "Optimization BESS",
            "Input Echo",
        ]

    def test_errors_tab_is_last(self, tmp_path: Path) -> None:
        rows = [
            _make_row_result(row_index=1, name="Good", analysis_type="buildability", results=_buildability_results()),
            _make_row_result(row_index=2, name="Bad", analysis_type="production", status="failed", error="err"),
        ]
        br = _make_batch_result(rows, tmp_path)
        input_rows = [_buildability_input_row(), _production_input_row()]
        path = generate_batch_workbook(br, input_rows)

        wb = load_workbook(path)
        assert wb.sheetnames[-1] == "Errors"
        assert wb.sheetnames[-2] == "Input Echo"

    def test_input_echo_before_errors(self, tmp_path: Path) -> None:
        """Input Echo is second-to-last when Errors tab exists."""
        rows = [
            _make_row_result(row_index=1, name="F", analysis_type="buildability", status="failed", error="err"),
        ]
        br = _make_batch_result(rows, tmp_path)
        path = generate_batch_workbook(br, [_buildability_input_row()])

        wb = load_workbook(path)
        assert wb.sheetnames == ["Buildability", "Input Echo", "Errors"]


# ── Tests: optimization BESS tab ───────────────────────────────────────


class TestOptimizationBessTab:
    """Verify Optimization BESS tab includes all expected columns."""

    def test_bess_specific_columns(self, tmp_path: Path) -> None:
        rr = _make_row_result(
            name="Site D",
            analysis_type="optimization_bess",
            results=_optimization_bess_results(),
        )
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_bess_input_row()])

        wb = load_workbook(path)
        ws = wb["Optimization BESS"]
        headers = [
            ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
        ]
        # Should have both optimization and BESS columns
        assert "Winner GCR" in headers
        assert "Winner NPV ($)" in headers
        assert "BESS Adds Value" in headers
        assert "BESS Power (MW)" in headers
        assert "BESS NPV ($)" in headers
        assert "Combined NPV ($)" in headers

    def test_bess_data_values(self, tmp_path: Path) -> None:
        rr = _make_row_result(
            name="Site D",
            analysis_type="optimization_bess",
            results=_optimization_bess_results(),
        )
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_bess_input_row()])

        wb = load_workbook(path)
        ws = wb["Optimization BESS"]
        col_map = {
            ws.cell(row=1, column=c).value: c
            for c in range(1, ws.max_column + 1)
        }
        assert ws.cell(row=2, column=col_map["BESS Adds Value"]).value is True
        assert ws.cell(row=2, column=col_map["BESS Power (MW)"]).value == 2.0
        assert ws.cell(row=2, column=col_map["Combined NPV ($)"]).value == 3200000.0


# ── Tests: column width ────────────────────────────────────────���──────


class TestColumnWidth:
    """Verify column widths are within bounds."""

    def test_column_widths_in_range(self, tmp_path: Path) -> None:
        rr = _make_row_result(results=_buildability_results())
        br = _make_batch_result([rr], tmp_path)
        path = generate_batch_workbook(br, [_buildability_input_row()])

        wb = load_workbook(path)
        ws = wb["Buildability"]
        for col_idx in range(1, ws.max_column + 1):
            letter = chr(64 + col_idx) if col_idx <= 26 else None
            if letter and letter in ws.column_dimensions:
                width = ws.column_dimensions[letter].width
                if width is not None:
                    assert width >= 10, f"Column {letter} too narrow: {width}"
                    assert width <= 40, f"Column {letter} too wide: {width}"
